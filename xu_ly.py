from models import HocSinh
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import calendar
import datetime
from file_manager import doc_diem_danh
from openpyxl import Workbook
from openpyxl.styles import PatternFill
def clean_filename(filename):
    """Remove invalid characters from filename"""
    return "".join(c for c in filename if c.isalnum() or c in (' ', '_', '-'))

def xuat_diem_danh_excel(ma_lop, ten_lop, danh_sach_hs, thang, nam):
    wb = Workbook()
    sheet = wb.active
    sheet.title = f"Diem Danh {clean_filename(ten_lop)} - Thang {thang}-{nam}"

    # Định dạng chung
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Định dạng tiêu đề
    sheet.merge_cells('A1:D1')
    title_cell = sheet['A1']
    title_cell.value = f"PHIẾU ĐIỂM DANH LỚP {ten_lop.upper()} - THÁNG {thang}/{nam}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = center_alignment
    title_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Định dạng header
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    headers = ["STT", "Họ và tên"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Tìm số ngày trong tháng và thêm vào header
    _, days_in_month = calendar.monthrange(nam, thang)
    for day in range(1, days_in_month + 1):
        col = day + 2
        cell = sheet.cell(row=3, column=col, value=day)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Cột tổng số buổi đi học
    total_col = days_in_month + 3
    cell = sheet.cell(row=3, column=total_col, value="Tổng số buổi đi học")
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border
    cell.fill = header_fill

    # Đọc dữ liệu điểm danh
    diem_danh_data = doc_diem_danh(ma_lop, thang, nam)

    # Ghi dữ liệu học sinh và điểm danh
    for row, hoc_sinh in enumerate(danh_sach_hs, start=4):
        sheet.cell(row=row, column=1, value=row-3).border = thin_border
        sheet.cell(row=row, column=2, value=hoc_sinh.ten_hs).border = thin_border

        total_days_present = 0
        for day in range(1, days_in_month + 1):
            cell = sheet.cell(row=row, column=day + 2)
            ngay = f"{nam}-{thang:02d}-{day:02d}"
            diem_danh_ngay = diem_danh_data.get(ngay, {}).get(hoc_sinh.ma_hs, "")
            if diem_danh_ngay == "Có mặt":
                cell.value = "X"
                cell.font = Font(color="00008B")  # Dark Blue
                total_days_present += 1
            cell.alignment = center_alignment
            cell.border = thin_border

        # Ghi tổng số buổi đi học
        total_cell = sheet.cell(row=row, column=total_col, value=total_days_present)
        total_cell.border = thin_border
        total_cell.alignment = center_alignment

    # Tự động điều chỉnh độ rộng cột
    for column in range(1, sheet.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(column)
        for cell in sheet[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    # Lưu file Excel
    ten_file = f"diem_danh_{clean_filename(ten_lop)}_{thang}_{nam}.xlsx"
    try:
        wb.save(ten_file)
        print(f"Đã xuất điểm danh sang file: {ten_file}")
    except PermissionError:
        print(f"Không có quyền ghi file '{ten_file}'.")
    except Exception as e:
        print(f"Lỗi khi xuất file Excel: {e}")